"""Tests del exportador determinista a Excel (`scripts/export_excel.py`)."""
import importlib.util
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent.parent
_spec = importlib.util.spec_from_file_location("export_excel", ROOT / "scripts" / "export_excel.py")
export_excel = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(export_excel)


def _df():
    return pd.DataFrame([
        {"grupo": "A", "fecha": "2026-06-11", "partido": "Mexico vs South Africa",
         "pick": "2-0", "modal": "1-0", "resultado": "Mexico", "E_pts": 1.14},
        {"grupo": "A", "fecha": "2026-06-11", "partido": "South Korea vs Czech Republic",
         "pick": "1-0", "modal": "1-1", "resultado": "South Korea", "E_pts": 0.78},
    ])


def test_write_predictions_modo_tabla(tmp_path):
    out = tmp_path / "pred.xlsx"
    export_excel.write_predictions(_df(), out_path=out)
    assert out.exists()
    ws = load_workbook(out)["Predicciones"]
    assert [c.value for c in ws[1]] == export_excel.HEADERS
    fila2 = [c.value for c in ws[2]]
    assert fila2[2] == "Mexico vs South Africa" and fila2[5] == "2-0"
    assert fila2[3] == "Mexico" and fila2[4] == "South Africa"   # local / visitante separados
    assert ws.max_row == 3   # cabecera + 2 partidos


def test_write_predictions_mapa_celdas_conserva_plantilla(tmp_path, monkeypatch):
    tpl = tmp_path / "plantilla.xlsx"
    wb = Workbook(); wb.active.title = "Grupos"; wb.active["A1"] = "mi formato"; wb.save(tpl)
    monkeypatch.setattr(export_excel, "EXCEL_MAP",
                        {"Mexico vs South Africa": ("Grupos", "C5")})
    out = tmp_path / "out.xlsx"
    export_excel.write_predictions(_df(), out_path=out, template=tpl)
    wb2 = load_workbook(out)
    assert wb2["Grupos"]["C5"].value == "2-0"      # marcador en su celda
    assert wb2["Grupos"]["A1"].value == "mi formato"  # se conserva el resto de la plantilla
