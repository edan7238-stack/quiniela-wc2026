"""Exporta las predicciones de grupos a Excel — DETERMINISTA (sin IA).

Toma la recomendación de fase de grupos del modelo (`quiniela.recommend_group_stage`) y la
escribe en una hoja de Excel. Dos modos:

  - **Plantilla** (`--excel ruta.xlsx`): abre tu archivo conservando su formato y vuelca la tabla
    de predicciones en una hoja `Predicciones` (crea o reemplaza esa hoja). El resto de tus hojas
    quedan intactas.
  - **Limpio** (sin `--excel`): genera un `Predicciones.xlsx` nuevo con la tabla.

Para mapear cada predicción a las CELDAS EXACTAS de TU plantilla (en vez de a una hoja nueva),
rellena `EXCEL_MAP` más abajo: `{"Local vs Visitante": ("Hoja", "C5"), ...}`. Mientras esté vacío
se usa el volcado en tabla.

Uso:
    python scripts/export_excel.py                       # genera Predicciones.xlsx
    python scripts/export_excel.py --excel quiniela.xlsx # rellena hoja "Predicciones" en tu archivo
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pandas as pd
from openpyxl import Workbook, load_workbook

# --------------------------------------------------------------------------- #
# Mapeo opcional a celdas concretas de TU plantilla. Vacío => volcado en tabla.
# Clave = texto del partido tal cual sale en la columna `partido` ("Local vs Visitante").
# Valor = (nombre_de_hoja, celda_destino_del_marcador). Ej.:
#   "Mexico vs South Africa": ("Grupos", "D5"),
# --------------------------------------------------------------------------- #
EXCEL_MAP: dict[str, tuple[str, str]] = {}

SHEET = "Predicciones"
HEADERS = ["Fecha", "Grupo", "Partido", "Local", "Visitante",
           "Pronóstico", "Resultado", "E[pts]"]


def _split_partido(partido: str) -> tuple[str, str]:
    local, _, visitante = str(partido).partition(" vs ")
    return local.strip(), visitante.strip()


def load_predictions() -> pd.DataFrame:
    """Construye el Predictor y devuelve la tabla de recomendación de grupos."""
    from src import predict, quiniela
    P = predict.Predictor()
    return quiniela.recommend_group_stage(P)["matches"]


def write_predictions(df: pd.DataFrame, out_path: Path, template: Path | None = None,
                      sheet: str = SHEET) -> Path:
    """Escribe la tabla `df` (salida de recommend_group_stage) en un Excel y guarda en `out_path`.

    Si `template` existe, se abre conservando el resto de hojas. Si `EXCEL_MAP` tiene entradas,
    coloca cada `pick` en su celda; si no, vuelca una tabla en la hoja `sheet`.
    """
    if template is not None and Path(template).exists():
        wb = load_workbook(template)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if EXCEL_MAP:
        # Modo celda-por-celda sobre la plantilla del usuario.
        for _, r in df.iterrows():
            dest = EXCEL_MAP.get(str(r["partido"]))
            if dest is None:
                continue
            hoja, celda = dest
            ws = wb[hoja] if hoja in wb.sheetnames else wb.create_sheet(hoja)
            ws[celda] = r["pick"]
    else:
        # Modo tabla: hoja `sheet` (reemplaza si ya existía).
        if sheet in wb.sheetnames:
            del wb[sheet]
        ws = wb.create_sheet(sheet)
        ws.append(HEADERS)
        has_fecha = "fecha" in df.columns
        for _, r in df.iterrows():
            local, visitante = _split_partido(r["partido"])
            ws.append([
                r.get("fecha", "") if has_fecha else "",
                r["grupo"], r["partido"], local, visitante,
                r["pick"], r["resultado"], float(r["E_pts"]),
            ])

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main() -> None:
    ap = argparse.ArgumentParser(description="Exporta predicciones de grupos a Excel.")
    ap.add_argument("--excel", type=Path, default=None,
                    help="Tu plantilla .xlsx (se rellena conservando el resto de hojas).")
    ap.add_argument("--out", type=Path, default=None,
                    help="Archivo de salida (def.: la plantilla, o ./Predicciones.xlsx).")
    args = ap.parse_args()

    out = args.out or args.excel or Path("Predicciones.xlsx")
    print("Calculando predicciones de grupos (esto tarda ~20 s)...")
    df = load_predictions()
    written = write_predictions(df, out_path=out, template=args.excel)
    print(f"Listo: {len(df)} partidos escritos en {written}")
    if not EXCEL_MAP and args.excel:
        print("Nota: se volcó la tabla en la hoja 'Predicciones'. Para colocar cada marcador en "
              "celdas concretas de tu plantilla, rellena EXCEL_MAP en este script.")


if __name__ == "__main__":
    main()
